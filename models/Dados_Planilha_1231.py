from sys import displayhook
import pandas
import Conexao
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.drawing.image import Image

# Busca de dados no banco
try:
    class Estacao1231():
        def _init_(self, data1, data2, Nome_Arquivo, Nome_Salvar):
            data2 = data2 + " 23:59:59"
            consulta_sql = "SELECT DATE(HoraLocal), SUM(Pluvio1h), SUM(Evaporacao1h) FROM medicoes WHERE Codigo_Sec = 1231 AND HoraLocal BETWEEN %s AND %s GROUP BY DATE(HoraLocal);"
            cursor = Conexao.con.cursor()
            cursor.execute(consulta_sql, (data1, data2))
            Dados = cursor.fetchall()

# Gerar dataframe com os dados
            df = pandas.DataFrame(
                Dados, columns=["DATA", "PRECIP. (mm)", "EVAPORACAO (mm)"])

# Carregar arquivo excel existente
            wb = load_workbook(Nome_Arquivo + ".xlsx")
            ws = wb["Evapo 1"]

# Leitura de parâmetros do arquivo
            row = ws.max_row
            column = ws.max_column

# Ler último valor acumulado
            Calculado = ws.cell(row=row, column=3).value
            Calculado = float(Calculado)  # type: ignore

# Condicao acumulativa da precipitacao mensal e insercao do valor no dataframe
            Acumulativo = df["PRECIP. (mm)"].cumsum() + Calculado
            df.insert(2, "PREC. ACUM. (mm)", Acumulativo)

# Formatacao da data
            df["DATA"] = pandas.to_datetime(df.DATA)
            # Ano com Y maiúsculo, saída com 4 dígitos / Ano com y minúsculo, saída com 2 dígitos
            df["DATA"] = df["DATA"].dt.strftime("%d/%m/%y")

# Transformar dataframe em datarows (linhas de dados)
            dr = dataframe_to_rows(df, index=False, header=False)

# Inserir dados na planilha
            for r in dr:
                ws.append(r)

# Inserir imagens na planilha
#            img1 = Image("C:/Users/Jair/Pictures/Acqua.png")
#            img2 = Image("C:/Users/Jair/Pictures/Lundin.png")

#            ws.add_image(img1, "A1")
#            ws.add_image(img2, "D1")

# Leitura de parâmetros do arquivo
            row_before = row + 1
            row_after = ws.max_row + 1

# Formatar dados da planilha
            for i in range(row_before, row_after):
                for j in range(1, column):
                    ws.cell(i, j).font = Font(name="Calibri",
                                              size=12)
#                                             bold = False,
#                                             italic = False,
                    ws.cell(i, j).border = Border(left=Side(border_style="thin",
                                                            color='FF000000'),
                                                  right=Side(border_style="thin",
                                                             color='FF000000'),
                                                  top=Side(border_style="thin",
                                                           color='FF000000'),
                                                  bottom=Side(border_style="thin",
                                                              color='FF000000'))
#                                                 diagonal=Side(border_style=None,
#                                                 color='FF000000'),
#                                                 diagonal_direction=0,
#                                                 outline=Side(border_style=None,
#                                                 color='FF000000'),
#                                                 vertical=Side(border_style=None,
#                                                 color='FF000000'),
#                                                 horizontal=Side(border_style=None,
#                                                 color='FF000000'))

                    ws.cell(i, j).alignment = Alignment(
                        horizontal='center', vertical='center')
                    ws.cell(i, j).number_format = '0.0'

# Apresentacao dos dataframes no terminal
#               displayhook(df)

# Exportar dataframes como arquivo xlsx
#               df.to_excel("Teste Salvamento.xlsx", index= False) # Gerar arquivo pelo pandas
            wb.save(Nome_Salvar + ".xlsx")  # Gerar arquivo pelo openpyxl
            print("\nArquivo excel criado com sucesso!!!\n")

except OSError as e:
    print("Erro: ", e)
